"""
Laporan Module - Reporting
"""
from pathlib import Path
from datetime import datetime, date, timedelta

import io

from flask import Blueprint, render_template, session, redirect, url_for, request, flash, send_file
from app.utils.auth import login_required, require_role
from app.models import Transaksi, Pelanggan, Pembayaran
from app import db
from app.transaksi.services import TransaksiService

BASE_DIR = Path(__file__).resolve().parent.parent
laporan_bp = Blueprint('laporan', __name__, template_folder=str(BASE_DIR / 'templates' / 'laporan'))


def format_currency(amount):
    try:
        return f"Rp {amount:,.0f}".replace(',', '.')
    except Exception:
        return 'Rp 0'


def get_date_range(filter_name, custom_date=None, custom_month=None):
    today = date.today()
    start = datetime(today.year, today.month, today.day)
    end = start + timedelta(days=1)

    if filter_name == 'hari_ini':
        return start, end

    if filter_name == 'kemarin':
        start = start - timedelta(days=1)
        end = start + timedelta(days=1)
        return start, end

    if filter_name == '7_hari':
        start = start - timedelta(days=6)
        return start, end

    if filter_name == '30_hari':
        start = start - timedelta(days=29)
        return start, end

    if filter_name == 'bulan_ini':
        start = datetime(today.year, today.month, 1)
        if today.month == 12:
            end = datetime(today.year + 1, 1, 1)
        else:
            end = datetime(today.year, today.month + 1, 1)
        return start, end

    if filter_name == 'per_tanggal' and custom_date:
        try:
            selected_date = date.fromisoformat(custom_date)
            start = datetime(selected_date.year, selected_date.month, selected_date.day)
            end = start + timedelta(days=1)
            return start, end
        except ValueError:
            return start, end

    if filter_name == 'per_bulan' and custom_month:
        try:
            selected_date = datetime.strptime(custom_month, '%Y-%m')
            start = datetime(selected_date.year, selected_date.month, 1)
            if selected_date.month == 12:
                end = datetime(selected_date.year + 1, 1, 1)
            else:
                end = datetime(selected_date.year, selected_date.month + 1, 1)
            return start, end
        except ValueError:
            return start, end

    return start, end


@laporan_bp.route('/')
def index():
    """Reports dashboard"""
    if login_required():
        return login_required()

    access = require_role('Admin', 'Kasir')
    if access:
        return access

    filter_range = request.args.get('range', 'hari_ini')
    custom_date = request.args.get('custom_date', '')
    custom_month = request.args.get('custom_month', '')

    start_date, end_date = get_date_range(filter_range, custom_date, custom_month)

    status_query = db.session.query(
        Transaksi.status_proses,
        db.func.count(Transaksi.id_transaksi)
    ).filter(
        Transaksi.is_active == True,
        Transaksi.tanggal_masuk >= start_date,
        Transaksi.tanggal_masuk < end_date
    ).group_by(Transaksi.status_proses).all()
    status_counts = {row[0]: row[1] for row in status_query}

    filtered_transaksi = Transaksi.query.filter(
        Transaksi.is_active == True,
        Transaksi.tanggal_masuk >= start_date,
        Transaksi.tanggal_masuk < end_date
    )

    total_transactions = filtered_transaksi.count()
    total_revenue = db.session.query(db.func.coalesce(db.func.sum(Pembayaran.jumlah), 0)).filter(
        Pembayaran.tanggal_pembayaran >= start_date,
        Pembayaran.tanggal_pembayaran < end_date
    ).scalar() or 0
    total_customers = db.session.query(db.func.count(db.distinct(Transaksi.id_pelanggan))).filter(
        Transaksi.is_active == True,
        Transaksi.tanggal_masuk >= start_date,
        Transaksi.tanggal_masuk < end_date
    ).scalar() or 0
    total_completed = status_counts.get('Selesai', 0)
    total_antrian = status_counts.get('Antrian', 0)

    recent_transactions = filtered_transaksi.order_by(Transaksi.tanggal_masuk.desc()).limit(8).all()

    filter_labels = {
        'hari_ini': 'Hari Ini',
        'kemarin': 'Kemarin',
        '7_hari': '7 Hari Terakhir',
        '30_hari': '30 Hari Terakhir',
        'bulan_ini': 'Bulan Ini',
        'per_tanggal': 'Per Tanggal',
        'per_bulan': 'Per Bulan'
    }

    return render_template(
        'laporan.html',
        status_counts=status_counts,
        recent_transactions=recent_transactions,
        total_transactions=total_transactions,
        total_revenue=format_currency(total_revenue),
        total_customers=total_customers,
        total_completed=total_completed,
        total_antrian=total_antrian,
        format_currency=format_currency,
        active_page='laporan',
        selected_filter=filter_range,
        custom_date=custom_date,
        custom_month=custom_month,
        filter_labels=filter_labels,
        selected_filter_label=filter_labels.get(filter_range, 'Hari Ini')
    )


def _build_export_filename(report_format, start_date, end_date):
    date_label = start_date.strftime('%Y%m%d')
    if end_date and end_date.date() != start_date.date():
        date_label += f"_to_{(end_date - timedelta(days=1)).strftime('%Y%m%d')}"
    ext = 'pdf' if report_format == 'pdf' else 'xlsx'
    return f"laporan_{date_label}.{ext}"


@laporan_bp.route('/export')
def export():
    """Export laporan ke PDF atau Excel"""
    if login_required():
        return login_required()

    access = require_role('Admin', 'Kasir')
    if access:
        return access

    report_format = request.args.get('format', 'pdf').lower()
    filter_range = request.args.get('range', 'hari_ini')
    custom_date = request.args.get('custom_date', '')
    custom_month = request.args.get('custom_month', '')

    start_date, end_date = get_date_range(filter_range, custom_date, custom_month)

    transaksi_query = Transaksi.query.filter(
        Transaksi.is_active == True,
        Transaksi.tanggal_masuk >= start_date,
        Transaksi.tanggal_masuk < end_date
    ).order_by(Transaksi.tanggal_masuk.desc())

    transaksi_list = transaksi_query.all()

    rows = []
    for trx in transaksi_list:
        pelanggan = db.session.get(Pelanggan, trx.id_pelanggan)
        rows.append({
            'nomor': trx.nomor_transaksi,
            'pelanggan': pelanggan.nama if pelanggan else '-',
            'status': trx.status_proses,
            'tanggal_masuk': trx.tanggal_masuk.strftime('%Y-%m-%d %H:%M') if trx.tanggal_masuk else '-',
            'estimasi_selesai': trx.tanggal_selesai_estimasi.strftime('%Y-%m-%d %H:%M') if trx.tanggal_selesai_estimasi else '-',
            'selesai_aktual': trx.tanggal_selesai_aktual.strftime('%Y-%m-%d %H:%M') if trx.tanggal_selesai_aktual else '-',
            'total_harga': format_currency(trx.total_harga or 0),
            'catatan': trx.catatan or ''
        })

    filename = _build_export_filename(report_format, start_date, end_date)

    if report_format == 'excel':
        try:
            from openpyxl import Workbook
        except ImportError:
            flash('Library openpyxl belum terpasang. Install openpyxl untuk export Excel.', 'danger')
            return redirect(url_for('laporan.index', range=filter_range, custom_date=custom_date, custom_month=custom_month))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Laporan'

        headers = ['Nomor', 'Pelanggan', 'Status', 'Tanggal Masuk', 'Estimasi Selesai', 'Selesai Aktual', 'Total Harga', 'Catatan']
        sheet.append(headers)

        for row in rows:
            sheet.append([
                row['nomor'],
                row['pelanggan'],
                row['status'],
                row['tanggal_masuk'],
                row['estimasi_selesai'],
                row['selesai_aktual'],
                row['total_harga'],
                row['catatan']
            ])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    if report_format == 'pdf':
        try:
            from fpdf import FPDF
        except ImportError:
            flash('Library fpdf belum terpasang. Install fpdf untuk export PDF.', 'danger')
            return redirect(url_for('laporan.index', range=filter_range, custom_date=custom_date, custom_month=custom_month))

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Laporan Transaksi', ln=True, align='C')
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 8, f'Periode: {filter_range} {custom_date or custom_month}', ln=True)
        pdf.ln(4)

        headers = ['No', 'Nomor', 'Pelanggan', 'Status', 'Tanggal Masuk', 'Estimasi', 'Aktual', 'Total']
        col_widths = [10, 35, 40, 28, 30, 30, 30, 27]

        for index, header in enumerate(headers):
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(col_widths[index], 8, header, border=1, align='C')
        pdf.ln()

        for idx, row in enumerate(rows, start=1):
            pdf.set_font('Arial', '', 8)
            values = [
                str(idx),
                row['nomor'],
                row['pelanggan'],
                row['status'],
                row['tanggal_masuk'],
                row['estimasi_selesai'],
                row['selesai_aktual'],
                row['total_harga']
            ]
            for col_index, value in enumerate(values):
                text = str(value).replace('\n', ' ')
                if len(text) > 40:
                    text = text[:40] + '...'
                pdf.cell(col_widths[col_index], 6, text, border=1)
            pdf.ln()

        pdf_output = pdf.output(dest='S').encode('latin-1', 'replace')
        return send_file(
            io.BytesIO(pdf_output),
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )

    flash('Format export tidak dikenal. Pilih PDF atau Excel.', 'danger')
    return redirect(url_for('laporan.index', range=filter_range, custom_date=custom_date, custom_month=custom_month))
